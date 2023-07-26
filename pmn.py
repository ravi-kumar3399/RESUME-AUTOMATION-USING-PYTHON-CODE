import PyPDF2
import docx2txt
import os
import re
import pandas as pd
from PyPDF2 import PdfFileReader
from openpyxl.styles import PatternFill 
from openpyxl.styles import Border,Side
from openpyxl.utils import rows_from_range, cols_from_range
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from openpyxl import workbook
import openpyxl
#skills = ['Python', 'Java', 'SQL', 'JavaScript', 'HTML', 'CSS', 'Machine Learning', 'Data Analysis', 'Project Management','c','c++','embedded']
# Define regex patterns to extract phone number and email
phone_pattern = re.compile(r'[+91\s]?\d{10}|\d{3}-\d{3}-\d{4}|\d{3} \d{3} \d{4}|[+91\s]?\d{2} \d{3} \d{3}|\d{10}/\d{10}')
email_pattern = re.compile(r'[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}')
name_pattern = re.compile(r'([A-Z][a-z]+)\s([A-Z][a-z]+)\s([A-Z][a-z]+)*')
experience_pattern= re.compile(r'\b(\d+)\s+years?\b')
#skill_pattern = re.compile(r'\b(' + '|'.join(skills) + r')\b')

# Create an empty list to store the extracted information
info_list = []

# Loop through the .docx files in the folder
folder_path = input('/path/to/folders:')
df = pd.DataFrame()

for filename in os.listdir(folder_path):
    if filename.endswith(".docx"):
        # Extract the text from the file
        file_path = os.path.join(folder_path, filename)
        text = docx2txt.process(file_path)
  
        name = text.split("\n")[0]   # Extract the name from the text (assuming the name is the first line)
  
        phone_match = phone_pattern.search(text) # Extract the phone number from the text using regex
        if phone_match:
            phone = phone_match.group()
        else:
            phone = ''

        email_match = email_pattern.search(text) # Extract the email address from the text using regex
        if email_match:
            email = email_match.group()
        else:
            email = ''
            
        experience_match = experience_pattern.search(text)
        if experience_match:
            experience=experience_match.group()
        else:
            experience = ''
        skills=re.findall("Python|AWS|Embedded|Linux|DevOps|Cloud computing|Machine learning|Java|Python Full stack developer|Java Full stack developer",text)
        mylist = list(dict.fromkeys(skills)) # removing duplicates
        #print("Skills:",mylist)

        # Add the information to the list
        info_list.append([name, phone, email,experience, mylist])
    elif filename.endswith(".pdf"):
            file_path = os.path.join(folder_path, filename)
            with open(file_path, 'rb') as pdf_file:
                reader = PyPDF2.PdfReader(pdf_file)
                page = reader.pages[0]
                text = page.extract_text()
            # Extract the name from the text (assuming the name is the first line)
            name = text.split("\n")[0]
   
            # Extract the phone number from the text using regex
            phone_match = phone_pattern.search(text)
            if phone_match:
                phone = phone_match.group()
            else:
                phone = ''

            # Extract the email address from the text using regex
            email_match = email_pattern.search(text)
            if email_match:
                email = email_match.group()
            else:
                email = ''
            experience_match=experience_pattern.search(text)
            if experience_match:
                experience=experience_match.group()
            else:
                experience = ''
                
            skills=re.findall("Python|AWS|Embedded|Linux|DevOps|Cloud computing|Machine learning|Java|Python Full stack developer|Java Full stack developer|c|c++|oracle",text)
            mylist = list(dict.fromkeys(skills)) # removing duplicates
            #print("Skills:",mylist)

            # Add the information to the list
            info_list.append([name, phone, email,experience, mylist])

    # Convert the list to a Pandas DataFrame
df = pd.DataFrame(info_list, columns=['Name', 'Phone', 'Email','Experience','domain'])

# Save the DataFrame to an Excel file
excel_path = input("enter excel path path:")

df.to_excel(excel_path,index=False)



# Add a new column at the beginning of the worksheet

# Load the Excel file
#excel_path = input(excel_path)
workbook = openpyxl.load_workbook(excel_path)

# Select the first worksheet
worksheet = workbook.worksheets[0]
worksheet.insert_cols(1)

# Set the header for the new column
#worksheet.cell(row=1, column=1, value="S.No.")
worksheet.column_dimensions['B'].auto_fit = True
# Add serial numbers to each row
for i in range(2, worksheet.max_row + 1):
    worksheet.cell(row=i, column=1, value=i - 1)
#import openpyxl
# Autofit columns and rows
for column in worksheet.columns:
    max_length = 0
    column_letter = get_column_letter(column[0].column)
    for cell in column:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    adjusted_width = (max_length + 2)
    worksheet.column_dimensions[column_letter].width = adjusted_width
for row in worksheet.rows:
    max_height = 0
    for cell in row:
        try:
            cell_text = str(cell.value)
            cell_lines = cell_text.split(' ')
            for line in cell_lines:
                if len(line) > max_height:
                    max_height = len(line)
        except:
            pass
    adjusted_height = max_height + 2
    worksheet.row_dimensions[row[0].row].height = adjusted_height
    worksheet.column_dimensions['B'].auto_fit = True





# Set the border for the first column
border = Border(left=Side(border_style='medium', color='FF000000'))

for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row):
    for cell in row:
        if cell.column == 1:
            cell.border = border

# Set the background color and font color for the first column
fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
font = Font(color='FFFFFF', bold=True)

for row in worksheet.iter_rows(min_row=1, max_row=1):
    for cell in row:
        cell.fill = fill
        cell.font = font

# Save the Excel file
workbook.save(excel_path)





