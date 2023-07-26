import glob
import openpyxl
import pandas as pd
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border,Side
path="C:\\Users\\DELL\\Desktop\\excel sheets"
file_list=glob.glob(path+"/*xlsx")
excel_file=[]
for file in file_list:
    excel_file.append(pd.read_excel(file))
excel_merged=pd.concat(excel_file)
excel_merged.to_excel("merged_file.xlsx",index=False)
workbook = openpyxl.load_workbook("merged_file.xlsx")

# Select the first worksheet
worksheet = workbook.worksheets[0]
#worksheet.insert_cols(1)

# Set the header for the new column
worksheet.cell(row=1, column=1, value="S.No.")
worksheet.column_dimensions['B'].auto_fit = True
# Add serial numbers to each row
for i in range(2, worksheet.max_row + 1):
    worksheet.cell(row=i, column=1, value=i - 1)
#import openpyxl
# Autofit columns and rows
for column in worksheet.columns:
    max_length = 0
    column_letter = get_column_letter(column[0].column)
    for cell in column:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    adjusted_width = (max_length + 2)
    worksheet.column_dimensions[column_letter].width = adjusted_width
for row in worksheet.rows:
    max_height = 0
    for cell in row:
        try:
            cell_text = str(cell.value)
            cell_lines = cell_text.split(' ')
            for line in cell_lines:
                if len(line) > max_height:
                    max_height = len(line)
        except:
            pass
    adjusted_height = max_height + 2
    worksheet.row_dimensions[row[0].row].height = adjusted_height
    worksheet.column_dimensions['B'].auto_fit = True


# Set the border for the first column
border = Border(left=Side(border_style='medium', color='FF000000'))

for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row):
    for cell in row:
        if cell.column == 1:
            cell.border = border

# Set the background color and font color for the first column
fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
font = Font(color='FFFFFF', bold=True)

for row in worksheet.iter_rows(min_row=1, max_row=1):
    for cell in row:
        cell.fill = fill
        cell.font = font

# Save the Excel file
workbook.save("merged_file.xlsx")
